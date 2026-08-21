"""
Crawler for EMA's downloadable medicines Excel report.

Source page:
    https://www.ema.europa.eu/en/medicines/download-medicine-data

The report is generated automatically by EMA and contains one row per medicine
on the Medicine sheet. The current workbook has metadata in row 1, headers in
row 9, and data beginning in row 10. The crawler discovers the header row by
looking for the stable "EMA product number" and "Medicine URL" columns instead
of relying solely on a fixed row number.
"""

from __future__ import annotations

import logging
import os
import tempfile
import time
from datetime import date, datetime
from typing import Dict, Iterator, List, Optional, Tuple
from urllib.parse import urljoin

from bs4 import BeautifulSoup
import openpyxl
import requests

from app.config import MAX_RECORDS_PER_COUNTRY
from app.db import check_record_exists_by_json_field, save_drug_record
from app.storage import build_document_key, content_type_for_ext, upload_file

logger = logging.getLogger(__name__)

REPORT_URL = os.getenv(
    'EMA_MEDICINES_REPORT_URL',
    'https://www.ema.europa.eu/en/documents/report/medicines-output-medicines-report_en.xlsx',
)
REQUEST_TIMEOUT = int(os.getenv('EMA_MEDICINES_TIMEOUT', '120'))
PAGE_TIMEOUT = int(os.getenv('EMA_MEDICINES_PAGE_TIMEOUT', '60'))
REQUEST_DELAY = float(os.getenv('EMA_MEDICINES_REQUEST_DELAY', '3'))
PDF_RETRIES = int(os.getenv('EMA_MEDICINES_PDF_RETRIES', '5'))

_COLUMN_MAP = {
    'Category': 'category',
    'Name of medicine': 'medicine_name',
    'EMA product number': 'ema_product_number',
    'Medicine status': 'medicine_status',
    'Opinion status': 'opinion_status',
    'Latest procedure affecting product information': 'latest_procedure',
    'International non-proprietary name (INN) / common name': 'inn',
    'Active substance': 'active_substance',
    'Therapeutic area (MeSH)': 'therapeutic_area',
    'Species\n(veterinary)': 'species_veterinary',
    'Patient safety': 'patient_safety',
    'ATC code (human)': 'atc_code_human',
    'ATCvet code (veterinary)': 'atc_code_veterinary',
    'Pharmacotherapeutic group\n(human)': 'pharmacotherapeutic_group_human',
    'Pharmacotherapeutic group\n(veterinary)': 'pharmacotherapeutic_group_veterinary',
    'Therapeutic indication': 'therapeutic_indication',
    'Accelerated assessment': 'accelerated_assessment',
    'Additional monitoring': 'additional_monitoring',
    'Advanced therapy': 'advanced_therapy',
    'Biosimilar': 'biosimilar',
    'Conditional approval': 'conditional_approval',
    'Exceptional circumstances': 'exceptional_circumstances',
    'Generic': 'generic',
    'Orphan medicine': 'orphan_medicine',
    'PRIME: priority medicine': 'prime_priority_medicine',
    'Marketing authorisation developer / applicant / holder': 'marketing_authorisation_holder',
    'European Commission decision date': 'european_commission_decision_date',
    'Start of rolling review date': 'start_of_rolling_review_date',
    'Start of evaluation date': 'start_of_evaluation_date',
    'Opinion adopted date': 'opinion_adopted_date',
    'Withdrawal of application date': 'withdrawal_of_application_date',
    'Marketing authorisation date': 'marketing_authorisation_date',
    'Refusal of marketing authorisation date': 'refusal_of_marketing_authorisation_date',
    'Withdrawal / expiry / revocation / lapse of marketing authorisation date': 'withdrawal_expiry_revocation_date',
    'Suspension of marketing authorisation date': 'suspension_of_marketing_authorisation_date',
    'Revision number': 'revision_number',
    'First published date': 'first_published_date',
    'Last updated date': 'last_updated_date',
    'Medicine URL': 'medicine_url',
}


def _clean(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip() or None
    return value


def _download_report() -> str:
    response = requests.get(
        REPORT_URL,
        headers={'User-Agent': 'source-predicate/EMA-medicine-report-crawler'},
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    handle = tempfile.NamedTemporaryFile(prefix='ema_medicines_', suffix='.xlsx', delete=False)
    try:
        handle.write(response.content)
        return handle.name
    finally:
        handle.close()


def _find_header_row(ws) -> Tuple[int, Dict[int, str]]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        headers = {
            index + 1: str(value).strip()
            for index, value in enumerate(row)
            if value is not None and str(value).strip()
        }
        if 'EMA product number' in headers.values() and 'Medicine URL' in headers.values():
            return row_number, headers
    raise ValueError('EMA report did not contain the expected medicine headers')


def _iter_rows(path: str) -> Iterator[Tuple[int, dict]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if 'Medicine' not in workbook.sheetnames:
            raise ValueError(f'EMA report has no Medicine sheet: {workbook.sheetnames}')
        ws = workbook['Medicine']
        header_row, headers = _find_header_row(ws)
        mapped = {
            column_number: _COLUMN_MAP[header]
            for column_number, header in headers.items()
            if header in _COLUMN_MAP
        }

        for row_number, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            header_row + 1,
        ):
            record = {}
            for column_number, field in mapped.items():
                if column_number <= len(values):
                    value = _clean(values[column_number - 1])
                    if value is not None:
                        record[field] = value
            if record.get('ema_product_number') and record.get('medicine_name'):
                yield row_number, record
    finally:
        workbook.close()


class EuropeanUnionMedicineExcelCrawler:
    """Ingests the official EMA medicines Excel report."""

    def __init__(self):
        self._report_path: Optional[str] = None
        self._session = requests.Session()
        self._session.headers.update({
            'User-Agent': 'source-predicate/EMA-medicine-crawler',
            'Accept-Language': 'en-US,en;q=0.9',
        })
        self._last_request_at = 0.0

    def close(self):
        if self._report_path:
            try:
                os.unlink(self._report_path)
            except FileNotFoundError:
                pass
            self._report_path = None
        self._session.close()

    def _throttle(self):
        elapsed = time.monotonic() - self._last_request_at
        if elapsed < REQUEST_DELAY:
            time.sleep(REQUEST_DELAY - elapsed)
        self._last_request_at = time.monotonic()

    @staticmethod
    def _find_english_pdf(section) -> Optional[str]:
        for link in section.select('a[href]'):
            href = link.get('href', '').strip()
            if href.lower().endswith('_en.pdf'):
                return href
        return None

    def _find_document_pdf(self, medicine_url: str) -> Tuple[Optional[str], Optional[str]]:
        """Return the English Product information PDF URL if present, otherwise the
        first available English PDF on the page (in document order), along with the
        EMA document type it was found under."""
        self._throttle()
        response = self._session.get(medicine_url, timeout=PAGE_TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'lxml')

        pi_section = soup.select_one('[data-ema-document-type="product-information"]')
        if pi_section:
            href = self._find_english_pdf(pi_section)
            if href:
                return urljoin(medicine_url, href), 'product_information'

        for section in soup.select('[data-ema-document-type]'):
            document_type = section.get('data-ema-document-type')
            if document_type == 'product-information':
                continue
            href = self._find_english_pdf(section)
            if href:
                return urljoin(medicine_url, href), document_type
        return None, None

    def _download_product_information(
        self,
        medicine_url: Optional[str],
        medicine_name: str,
        country_id: int,
    ) -> Tuple[Optional[str], List[dict]]:
        if not medicine_url:
            return None, []

        try:
            pdf_url, document_type = self._find_document_pdf(medicine_url)
        except requests.RequestException as exc:
            logger.warning('EMA medicine page failed for %s: %s', medicine_url, exc)
            return None, []
        if not pdf_url:
            logger.info('No English PDF document found for %s', medicine_name)
            return None, []
        if document_type != 'product_information':
            logger.info(
                'No Product information PDF for %s; falling back to %s document',
                medicine_name, document_type,
            )

        content = None
        content_type = ''
        status = 0
        for attempt in range(1, PDF_RETRIES + 1):
            self._throttle()
            try:
                response = self._session.get(
                    pdf_url,
                    headers={'Accept': 'application/pdf'},
                    timeout=PAGE_TIMEOUT,
                )
                status = response.status_code
                content_type = response.headers.get('Content-Type', '')
                if status == 200:
                    content = response.content
                    break
                if status == 429:
                    retry_after = response.headers.get('Retry-After')
                    try:
                        wait = max(1, int(float(retry_after))) if retry_after else 60
                    except ValueError:
                        wait = 60
                    logger.warning(
                        'EMA rate limited %s; waiting %ss before retry %s/%s',
                        pdf_url, wait, attempt, PDF_RETRIES,
                    )
                    time.sleep(wait)
                    continue
                if 500 <= status < 600:
                    time.sleep(min(60, 2 ** attempt))
                    continue
                break
            except requests.RequestException as exc:
                logger.warning(
                    'Product information PDF request failed for %s (attempt %s/%s): %s',
                    medicine_name, attempt, PDF_RETRIES, exc,
                )
                time.sleep(min(60, 2 ** attempt))

        if not content or status != 200:
            logger.warning('Product information PDF failed for %s: HTTP %s', medicine_name, status)
            return None, []
        if 'pdf' not in (content_type or '').lower() and not pdf_url.lower().endswith('.pdf'):
            logger.warning('Skipping non-PDF Product information response for %s', medicine_name)
            return None, []

        key = build_document_key(f'european_union/{country_id}', medicine_name, '.pdf')
        s3_path = upload_file(content, key, content_type_for_ext('.pdf'))
        return s3_path, [{
            'document_type': document_type,
            'language': 'en',
            'source_url': pdf_url,
            's3_path': s3_path,
            'content_type': content_type,
            'size_bytes': len(content),
        }]

    def process_country(self, country_id: int):
        self._report_path = _download_report()
        saved = 0
        seen = 0

        try:
            for row_number, record in _iter_rows(self._report_path):
                seen += 1
                product_number = record['ema_product_number']
                if check_record_exists_by_json_field(country_id, 'ema_product_number', product_number):
                    continue

                json_data = {
                    'source': 'ema_medicines_excel',
                    'report_url': REPORT_URL,
                    'report_sheet': 'Medicine',
                    'report_row': row_number,
                    **record,
                }
                s3_path, documents = self._download_product_information(
                    record.get('medicine_url'),
                    str(record['medicine_name']),
                    country_id,
                )
                json_data['documents'] = documents
                save_drug_record(
                    str(record['medicine_name'])[:255],
                    country_id,
                    [s3_path] if s3_path else None,
                    json_data,
                )
                saved += 1

                if MAX_RECORDS_PER_COUNTRY and saved >= MAX_RECORDS_PER_COUNTRY:
                    break
        finally:
            self.close()

        logger.info('EMA medicine Excel crawl finished: seen=%s saved=%s', seen, saved)
